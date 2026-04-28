import os
import io
import logging
from datetime import date, datetime, timezone

from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, abort
from apscheduler.schedulers.background import BackgroundScheduler
from dotenv import load_dotenv

load_dotenv()

import db
import sync
import tikhub

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# ── TEMPLATE FILTERS ──────────────────────────────────────────────────────────

@app.template_filter('fromjson')
def fmt_fromjson(s):
    if not s:
        return []
    try:
        import json as _json
        return _json.loads(s)
    except Exception:
        return []


@app.template_filter('num')
def fmt_num(n):
    if n is None:
        return '—'
    n = float(n)
    if n >= 1_000_000:
        v = n / 1_000_000
        return f'{v:.1f}M' if v % 1 != 0 else f'{int(v)}M'
    if n >= 1_000:
        v = n / 1_000
        return f'{v:.1f}K' if v % 1 != 0 else f'{int(v)}K'
    return str(int(n))


@app.template_filter('gmv')
def fmt_gmv(n):
    if not n:
        return '—'
    n = float(n)
    if n >= 1_000:
        return f'${n/1000:.1f}K'
    return f'${n:.0f}'


@app.template_filter('datefmt')
def fmt_date(d):
    if d is None:
        return '—'
    if isinstance(d, str):
        try:
            d = date.fromisoformat(d)
        except ValueError:
            return d
    return d.strftime('%b %-d')


@app.template_filter('timeago')
def fmt_timeago(dt):
    if dt is None:
        return 'never'
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt)
        except ValueError:
            return dt
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    diff = int((now - dt).total_seconds())
    if diff < 60:
        return 'just now'
    if diff < 3600:
        return f'{diff // 60}m ago'
    if diff < 86400:
        return f'{diff // 3600}h ago'
    return f'{diff // 86400}d ago'


@app.template_filter('duration')
def fmt_duration(secs):
    if secs is None:
        return ''
    secs = int(secs)
    m, s = divmod(secs, 60)
    return f'{m}:{s:02d}'


# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route('/')
def dashboard():
    clients = db.get_all_clients_with_period_stats()
    if not clients:
        return redirect(url_for('settings'))

    client_id = request.args.get('client', type=int)
    active = next((c for c in clients if c['id'] == client_id), clients[0])

    period_id = request.args.get('period', type=int)
    active_period = db.get_active_period(active['id'])
    all_periods = db.get_all_periods_for_client(active['id'])

    selected_period = None
    if period_id:
        selected_period = next((p for p in all_periods if p['id'] == period_id), None)
    if not selected_period:
        selected_period = active_period

    period_start = selected_period.get('period_start') if selected_period else None
    period_end = selected_period.get('period_end') if selected_period else None

    stats = db.get_client_stats(active['id'], period_start=period_start)
    filter_type = request.args.get('filter')
    videos = db.get_client_videos(
        active['id'],
        filter_type=filter_type,
        limit=30,
        period_start=period_start,
        period_end=period_end,
    )
    top_products = db.get_top_products(active['id'])
    recent = db.get_recent_activity(active['id'])
    products_info = db.get_products_info_map()
    creator_handle = db.get_setting('creator_handle', '')
    today = date.today()

    return render_template(
        'dashboard.html',
        clients=clients,
        active=active,
        stats=stats,
        videos=videos,
        top_products=top_products,
        recent=recent,
        active_period=active_period,
        selected_period=selected_period,
        all_periods=all_periods,
        today=today,
        filter_type=filter_type,
        products_info=products_info,
        creator_handle=creator_handle,
    )


@app.route('/products')
def products():
    client_id = request.args.get('client', type=int)
    all_products = db.get_all_products_with_stats()
    summary = db.get_product_summary_stats()
    clients = db.get_all_clients()

    if client_id:
        filtered = [p for p in all_products if p['client_id'] == client_id]
    else:
        filtered = all_products

    # Attach video strips to each product
    for p in filtered:
        p['video_strip'] = db.get_product_videos(p['product_id'])

    return render_template(
        'products.html',
        products=filtered,
        all_products=all_products,
        summary=summary,
        clients=clients,
        active_client_id=client_id,
    )


@app.route('/videos')
def videos():
    client_id = request.args.get('client', type=int)
    filter_type = request.args.get('filter')
    period_id = request.args.get('period', type=int)
    clients = db.get_all_clients()

    period_start = None
    period_end = None
    selected_period = None
    all_periods = []
    if client_id:
        all_periods = db.get_all_periods_for_client(client_id)
        if period_id:
            selected_period = next((p for p in all_periods if p['id'] == period_id), None)
        if selected_period:
            period_start = selected_period['period_start']
            period_end = selected_period['period_end']

    all_videos = db.get_all_videos(
        client_id=client_id,
        filter_type=filter_type,
        period_start=period_start,
        period_end=period_end,
    )
    creator_handle = db.get_setting('creator_handle', '')
    return render_template(
        'videos.html',
        videos=all_videos,
        clients=clients,
        active_client_id=client_id,
        filter_type=filter_type,
        selected_period=selected_period,
        all_periods=all_periods,
        creator_handle=creator_handle,
    )


@app.route('/settings')
def settings():
    clients = db.get_all_clients()
    client_data = []
    for c in clients:
        client_data.append({
            **c,
            'products': db.get_client_products(c['id']),
            'active_period': db.get_active_period(c['id']),
            'period_history': db.get_period_history(c['id']),
        })
    creator_handle = db.get_setting('creator_handle', '')
    return render_template('settings.html', client_data=client_data,
                           creator_handle=creator_handle, today=date.today())


# ── SETTINGS ACTIONS ──────────────────────────────────────────────────────────

@app.route('/settings/creator/update', methods=['POST'])
def update_creator():
    handle = request.form.get('creator_handle', '').strip().lstrip('@')
    db.set_setting('creator_handle', handle)
    return redirect(url_for('settings'))


@app.route('/settings/clients/add', methods=['POST'])
def add_client():
    brand_name = request.form['brand_name'].strip()
    tiktok_handle = request.form.get('tiktok_handle', '').strip().lstrip('@')
    brand_color = request.form.get('brand_color', '#ffffff')
    post_target = int(request.form.get('post_target', 30))
    db.add_client(brand_name, tiktok_handle, brand_color, post_target)
    return redirect(url_for('settings'))


@app.route('/settings/clients/<int:client_id>/update', methods=['POST'])
def update_client(client_id):
    brand_name = request.form['brand_name'].strip()
    tiktok_handle = request.form.get('tiktok_handle', '').strip().lstrip('@')
    brand_color = request.form.get('brand_color', '#ffffff')
    post_target = int(request.form.get('post_target', 30))
    db.update_client(client_id, brand_name, tiktok_handle, brand_color, post_target)
    return redirect(url_for('settings'))


@app.route('/settings/clients/<int:client_id>/delete', methods=['POST'])
def delete_client(client_id):
    db.delete_client(client_id)
    return redirect(url_for('settings'))


@app.route('/settings/products/add', methods=['POST'])
def add_product():
    client_id = int(request.form['client_id'])
    product_id = request.form['product_id'].strip()
    name, thumbnail_url = tikhub.lookup_product_info(product_id)
    product_name = name or f"Product {product_id}"
    db.add_product(client_id, product_id, product_name, thumbnail_url)
    return redirect(url_for('settings'))


@app.route('/settings/products/<int:product_db_id>/update', methods=['POST'])
def update_product(product_db_id):
    import base64 as _b64
    product_name = request.form.get('product_name', '').strip()
    thumbnail_url = request.form.get('thumbnail_url', '').strip()
    file = request.files.get('thumbnail_file')
    if file and file.filename:
        data = file.read()
        mime = file.content_type or 'image/jpeg'
        thumbnail_url = f"data:{mime};base64,{_b64.b64encode(data).decode()}"
    db.set_product_info(product_db_id, product_name, thumbnail_url or None)
    return redirect(url_for('settings'))


@app.route('/settings/products/<int:product_db_id>/delete', methods=['POST'])
def delete_product(product_db_id):
    db.delete_product(product_db_id)
    return redirect(url_for('settings'))


@app.route('/settings/periods/<int:client_id>/start', methods=['POST'])
def start_period(client_id):
    period_start = request.form.get('period_start') or str(date.today())
    db.start_period(client_id, period_start)
    return redirect(url_for('settings'))


@app.route('/settings/periods/<int:period_id>/update', methods=['POST'])
def update_period(period_id):
    period_start = request.form.get('period_start')
    period_end = request.form.get('period_end')
    db.update_period(period_id, period_start, period_end)
    return redirect(url_for('settings'))


@app.route('/settings/periods/<int:period_id>/complete', methods=['POST'])
def complete_period(period_id):
    db.complete_period(period_id)
    return redirect(url_for('settings'))


# ── REPORTS ───────────────────────────────────────────────────────────────────

@app.route('/reports')
def reports():
    clients = db.get_all_clients()
    client_id = request.args.get('client', type=int)
    periods = []
    if client_id:
        periods = db.get_all_periods_for_client(client_id)
    return render_template('reports.html', clients=clients, active_client_id=client_id, periods=periods)


@app.route('/reports/download')
def reports_download():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    client_id = request.args.get('client', type=int)
    period_id = request.args.get('period', type=int)
    if not client_id or not period_id:
        return redirect(url_for('reports'))

    client = db.get_client(client_id)
    period = db.get_period(period_id)
    if not client or not period:
        abort(404)

    summary = db.get_report_summary(client_id, period['period_start'], period['period_end'])
    vid_rows = db.get_report_videos(client_id, period['period_start'], period['period_end'])

    wb = openpyxl.Workbook()

    # ── SHEET 1: SUMMARY ──
    ws1 = wb.active
    ws1.title = 'Summary'

    accent = '3DDC84'
    black = '111111'
    light_gray = 'F6F6F6'
    mid_gray = 'E4E4E4'

    # Column widths
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 34

    def hdr_style(cell, text):
        cell.value = text
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill('solid', fgColor=black)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    def label_style(cell, text):
        cell.value = text
        cell.font = Font(name='Calibri', size=10, color='888888')
        cell.fill = PatternFill('solid', fgColor=light_gray)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    def value_style(cell, text, bold=False, green=False):
        cell.value = text
        color = accent if green else black
        cell.font = Font(name='Calibri', size=10, bold=bold, color=color)
        cell.fill = PatternFill('solid', fgColor='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    thin = Side(style='thin', color=mid_gray)
    border = Border(bottom=Side(style='thin', color=mid_gray))

    # Title row
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells('A1:B1')
    title_cell = ws1['A1']
    title_cell.value = f'TRAKR — Brand Report'
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor=black)
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    period_start_fmt = period['period_start'].strftime('%b %-d, %Y') if period['period_start'] else ''
    period_end_fmt = period['period_end'].strftime('%b %-d, %Y') if period['period_end'] else ''
    target = period.get('target_posts') or client.get('post_target') or 30
    videos_posted = summary['video_count'] or 0
    pct = round(min(videos_posted / target * 100, 100), 1) if target else 0

    rows = [
        ('Brand',              client['brand_name'],                          False, False),
        ('Report Period',      f"{period_start_fmt} – {period_end_fmt}",      False, False),
        ('Period Status',      (period.get('status') or '').title(),          False, False),
        ('Videos Posted',      videos_posted,                                 True,  False),
        ('Post Target',        target,                                        False, False),
        ('Completion',         f"{pct}%",                                     True,  True),
        ('Total Views',        summary['total_views'] or 0,                   False, False),
        ('Total Likes',        summary['total_likes'] or 0,                   False, False),
        ('Total Comments',     summary['total_comments'] or 0,                False, False),
        ('Total GMV',          f"${float(summary['total_gmv'] or 0):,.2f}",   True,  True),
        ('Total Orders',       summary['total_orders'] or 0,                  False, False),
        ('Report Generated',   datetime.now().strftime('%b %-d, %Y'),         False, False),
    ]

    for i, (label, value, bold, green) in enumerate(rows, start=2):
        ws1.row_dimensions[i].height = 22
        label_style(ws1.cell(row=i, column=1), label)
        value_style(ws1.cell(row=i, column=2), value, bold=bold, green=green)
        ws1.cell(row=i, column=1).border = border
        ws1.cell(row=i, column=2).border = border

    # ── SHEET 2: VIDEOS ──
    ws2 = wb.create_sheet('Videos')

    headers = ['Posted Date', 'Description', 'Duration', 'Views', 'Likes', 'Comments', 'GMV', 'Orders', 'Product', 'TikTok Link']
    col_widths = [14, 48, 10, 12, 12, 12, 12, 10, 28, 52]
    for j, (h, w) in enumerate(zip(headers, col_widths), start=1):
        ws2.column_dimensions[get_column_letter(j)].width = w
        cell = ws2.cell(row=1, column=j, value=h)
        cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.fill = PatternFill('solid', fgColor=black)
        cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    ws2.row_dimensions[1].height = 22
    ws2.freeze_panes = 'A2'

    creator_handle = db.get_setting('creator_handle', '')

    def duration_fmt(secs):
        if not secs:
            return ''
        secs = int(secs)
        m, s = divmod(secs, 60)
        return f'{m}:{s:02d}'

    alt_fill = PatternFill('solid', fgColor='FAFAFA')
    for row_idx, v in enumerate(vid_rows, start=2):
        ws2.row_dimensions[row_idx].height = 18
        fill = alt_fill if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')
        posted = v['posted_at'].strftime('%b %-d, %Y') if v.get('posted_at') else ''
        gmv_val = f"${float(v['gmv']):,.2f}" if v.get('gmv') else '—'
        tiktok_url = f"https://www.tiktok.com/@{creator_handle}/video/{v['video_id']}" if creator_handle else v['video_id']
        row_data = [
            posted,
            v.get('description') or '',
            duration_fmt(v.get('duration')),
            v.get('views') or 0,
            v.get('likes') or 0,
            v.get('comments') or 0,
            gmv_val,
            v.get('orders') or 0,
            v.get('product_name') or (v.get('tagged_product_id') or ''),
            tiktok_url,
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name='Calibri', size=10)
            cell.fill = fill
            cell.alignment = Alignment(vertical='center', indent=1)
            cell.border = Border(bottom=Side(style='thin', color='F2F2F2'))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = ''.join(c for c in client['brand_name'] if c.isalnum() or c in (' ', '-', '_')).strip()
    period_slug = f"{period['period_start']}_{period['period_end']}"
    filename = f"Trakr_{safe_name}_{period_slug}.xlsx"

    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── SHAREABLE BRAND PAGE ──────────────────────────────────────────────────────

@app.route('/brand/<token>')
def brand_share(token):
    client = db.get_client_by_token(token)
    if not client:
        abort(404)
    all_periods = db.get_all_periods_for_client(client['id'])
    period_id = request.args.get('period', type=int)
    selected_period = None
    if period_id:
        selected_period = next((p for p in all_periods if p['id'] == period_id), None)
    if not selected_period:
        selected_period = next((p for p in all_periods if p.get('status') == 'active'), None)
        if not selected_period and all_periods:
            selected_period = all_periods[0]

    period_start = selected_period['period_start'] if selected_period else None
    period_end = selected_period['period_end'] if selected_period else None

    stats = db.get_client_stats(client['id'], period_start=period_start)
    videos = db.get_client_videos(
        client['id'], limit=50,
        period_start=period_start, period_end=period_end,
    )
    products_info = db.get_products_info_map()
    creator_handle = db.get_setting('creator_handle', '')
    today = date.today()
    return render_template(
        'brand_share.html',
        client=client,
        stats=stats,
        videos=videos,
        selected_period=selected_period,
        all_periods=all_periods,
        today=today,
        products_info=products_info,
        creator_handle=creator_handle,
        token=token,
    )


# ── SYNC API ──────────────────────────────────────────────────────────────────

@app.route('/api/sync', methods=['POST'])
def trigger_sync_all():
    try:
        count = sync.sync_creator()
        return jsonify({'status': 'ok', 'videos_fetched': count})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sync/<int:client_id>', methods=['POST'])
def trigger_sync_client(client_id):
    try:
        count = sync.sync_creator()
        return jsonify({'status': 'ok', 'videos_fetched': count})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/sync/gmv', methods=['POST'])
def trigger_sync_gmv():
    try:
        count = sync.sync_gmv()
        return jsonify({'status': 'ok', 'videos_enriched': count})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/client-periods/<int:client_id>')
def api_client_periods(client_id):
    periods = db.get_all_periods_for_client(client_id)
    result = []
    for p in periods:
        result.append({
            'id': p['id'],
            'period_start': str(p['period_start']),
            'period_end': str(p['period_end']),
            'status': p['status'],
            'target_posts': p['target_posts'],
        })
    return jsonify(result)


@app.route('/api/clients/reorder', methods=['POST'])
def reorder_clients():
    data = request.get_json()
    ordered_ids = data.get('order', []) if data else []
    if not ordered_ids:
        return jsonify({'status': 'error', 'message': 'order required'}), 400
    db.update_client_sort_orders([int(i) for i in ordered_ids])
    return jsonify({'status': 'ok'})


@app.route('/settings/clients/<int:client_id>/token/regenerate', methods=['POST'])
def regenerate_token(client_id):
    db.regenerate_share_token(client_id)
    return redirect(url_for('settings'))


@app.route('/api/videos/<video_id>/assign', methods=['POST'])
def assign_video(video_id):
    data = request.get_json()
    client_id = data.get('client_id') if data else None
    if not client_id:
        return jsonify({'status': 'error', 'message': 'client_id required'}), 400
    db.assign_video_to_client(video_id, int(client_id))
    return jsonify({'status': 'ok'})


# ── STARTUP ───────────────────────────────────────────────────────────────────

def create_app():
    try:
        db.init_db()
    except Exception as e:
        logger.error(f"DB init failed at startup: {e}")
        logger.error("App will start but DB operations will fail — check DATABASE_URL")

    try:
        scheduler = BackgroundScheduler()
        scheduler.add_job(sync.sync_all, 'interval', hours=24, id='daily_sync')
        scheduler.start()
    except Exception as e:
        logger.error(f"Scheduler failed to start: {e}")

    return app


if __name__ == '__main__':
    create_app()
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
else:
    create_app()
